import json
import os
from datetime import datetime
from pathlib import Path
from urllib import parse, request, error

from openpyxl import load_workbook


def load_env(env_path: Path) -> None:
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def cell_text(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s


def cell_num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_invoice(xlsx_path: Path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=12, values_only=True)]

    client_name = ""
    invoice_number = ""
    issue_date = ""
    grand_total = 0.0
    payment_deadline = ""
    payment_schedule = []

    sec_titles = [
        "ELECTRICAL AND ELECTRONICS COMPONENTS PRICES",
        "PROTOTYPE ITEMS",
        "SERVICE COST",
    ]
    sections = [
        {"title": sec_titles[0], "columns": [{"key": "sn", "label": "S/N"}, {"key": "item", "label": "Item"}, {"key": "qty", "label": "Qnty"}, {"key": "unitPrice", "label": "Unit price"}, {"key": "totalPrice", "label": "Total price"}], "rows": []},
        {"title": sec_titles[1], "columns": [{"key": "sn", "label": "S/N"}, {"key": "item", "label": "Item"}, {"key": "qty", "label": "Qnty"}, {"key": "unitPrice", "label": "Unit price"}, {"key": "totalPrice", "label": "Total price"}], "rows": []},
        {"title": sec_titles[2], "columns": [{"key": "sn", "label": "S/N"}, {"key": "item", "label": "Item"}, {"key": "qty", "label": "Qnty"}, {"key": "amount", "label": "Amount"}], "rows": []},
    ]

    section_idx = -1
    in_schedule = False
    for r in rows:
        c0, c1, c3, c5, c7, c8 = cell_text(r[0]), cell_text(r[1]), cell_text(r[3]), cell_text(r[5]), cell_text(r[7]), cell_text(r[8])
        up0 = c0.upper()

        if c0 == "Bill to:":
            client_name = c1
        if c7 == "Invoice no:":
            invoice_number = c8
        if c7 == "Invoice Date:":
            issue_date = cell_text(r[8])

        if up0 in sec_titles:
            section_idx = sec_titles.index(up0)
            in_schedule = False
            continue

        if up0 == "PAYMENT SCHEDULE":
            in_schedule = True
            section_idx = -1
            continue

        if in_schedule:
            if isinstance(r[0], (int, float)):
                payment_schedule.append(
                    {
                        "phase": str(int(cell_num(r[0]))),
                        "description": cell_text(r[1]),
                        "amountRatio": cell_num(r[3]),
                        "amountToPay": cell_num(r[5]),
                        "deadline": cell_text(r[7]),
                    }
                )
            elif cell_text(r[1]).upper() == "FINAL":
                payment_deadline = cell_text(r[7])
            continue

        if c3.upper() == "GRAND TOTAL":
            grand_total = cell_num(r[7])
            continue

        if section_idx >= 0 and isinstance(r[0], (int, float)):
            sn = str(int(cell_num(r[0])))
            item = cell_text(r[1])
            qty = cell_num(r[3])
            if section_idx < 2:
                unit_price = cell_num(r[5])
                total_price = cell_num(r[7])
                sections[section_idx]["rows"].append(
                    {
                        "sn": sn,
                        "item": item,
                        "qty": f"{qty:g}",
                        "unitPrice": f"{unit_price:g}",
                        "totalPrice": f"{total_price:g}",
                    }
                )
            else:
                amount = cell_num(r[7])
                sections[section_idx]["rows"].append(
                    {
                        "sn": sn,
                        "item": item,
                        "qty": f"{qty:g}",
                        "amount": f"{amount:g}",
                    }
                )

    if not invoice_number:
        invoice_number = f"HC-PI-IMPORT-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    if not issue_date:
        issue_date = datetime.now().strftime("%Y-%m-%d")
    due_date = ""
    if payment_schedule:
        due_date = payment_schedule[-1]["deadline"]
    elif payment_deadline and payment_deadline[:4].isdigit():
        due_date = payment_deadline

    subtotal = grand_total
    payload = {
        "dashboardScope": "project",
        "invoiceNumber": invoice_number,
        "clientName": client_name,
        "issueDate": issue_date,
        "dueDate": due_date,
        "currency": "TZS",
        "totals": {
            "subtotal": subtotal,
            "taxAmount": 0,
            "grandTotal": grand_total,
        },
        "projectTables": {
            "sections": sections,
            "paymentSchedule": payment_schedule,
            "paymentGrandTotal": f"{grand_total:g}",
            "paymentDeadline": payment_deadline,
            "note": f"Imported from Excel: {xlsx_path.name}",
        },
    }

    return {
        "invoice_number": invoice_number,
        "client_name": client_name or "Unknown Client",
        "issue_date": issue_date,
        "due_date": due_date or None,
        "currency": "TZS",
        "subtotal": subtotal,
        "tax_amount": 0,
        "grand_total": grand_total,
        "created_by": None,
        "payload": payload,
    }


def supabase_request(method: str, url: str, apikey: str, body=None):
    headers = {
        "apikey": apikey,
        "Authorization": f"Bearer {apikey}",
    }
    data = None
    if body is not None:
        headers["Content-Type"] = "application/json"
        headers["Prefer"] = "return=representation"
        data = json.dumps(body).encode("utf-8")
    req = request.Request(url, method=method, headers=headers, data=data)
    try:
        with request.urlopen(req) as resp:
            raw = resp.read().decode("utf-8")
            return resp.status, json.loads(raw) if raw else None
    except error.HTTPError as e:
        raw = e.read().decode("utf-8")
        try:
            payload = json.loads(raw)
        except Exception:
            payload = {"error": raw}
        return e.code, payload


def ensure_unique_invoice_number(base_url: str, apikey: str, invoice_number: str) -> str:
    safe = parse.quote(invoice_number, safe="")
    status, data = supabase_request("GET", f"{base_url}/rest/v1/invoices?select=id&invoice_number=eq.{safe}&limit=1", apikey)
    if status >= 400:
        return invoice_number
    if isinstance(data, list) and data:
        return f"{invoice_number}-IMP"
    return invoice_number


def main():
    repo_root = Path(__file__).resolve().parents[1]
    load_env(repo_root / ".env.local")

    supabase_url = os.getenv("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
    apikey = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("NEXT_PUBLIC_SUPABASE_ANON_KEY")
    if not supabase_url or not apikey:
        raise SystemExit("Missing Supabase env vars in .env.local")

    xlsx = Path(r"J:\My Drive\Financial records 2026\Proforma invoice 2026\Zakaria profoma invoice 11.xlsx")
    record = parse_invoice(xlsx)
    record["invoice_number"] = ensure_unique_invoice_number(supabase_url, apikey, record["invoice_number"])
    record["payload"]["invoiceNumber"] = record["invoice_number"]

    now = datetime.utcnow().isoformat()
    record["created_at"] = now
    record["updated_at"] = now

    status, data = supabase_request("POST", f"{supabase_url}/rest/v1/invoices", apikey, record)
    if status >= 400:
        raise SystemExit(f"Insert failed ({status}): {json.dumps(data, ensure_ascii=False)}")

    first = data[0] if isinstance(data, list) and data else {}
    print(
        json.dumps(
            {
                "inserted": True,
                "id": first.get("id"),
                "invoice_number": first.get("invoice_number"),
                "client_name": first.get("client_name"),
                "grand_total": first.get("grand_total"),
                "source_file": str(xlsx),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()

