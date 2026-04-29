import json
import os
from datetime import UTC, datetime
from pathlib import Path
from urllib import error, parse, request

from openpyxl import load_workbook


def load_env(env_path: Path) -> None:
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def cell_text(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def cell_num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except Exception:
        return 0.0


def supabase_request(method: str, url: str, apikey: str, body=None):
    headers = {"apikey": apikey, "Authorization": f"Bearer {apikey}"}
    data = None
    if body is not None:
        headers["Content-Type"] = "application/json"
        headers["Prefer"] = "return=representation"
        data = json.dumps(body).encode("utf-8")
    req = request.Request(url, method=method, headers=headers, data=data)
    try:
        with request.urlopen(req) as resp:
            raw = resp.read().decode("utf-8")
            return resp.status, (json.loads(raw) if raw else None)
    except error.HTTPError as e:
        raw = e.read().decode("utf-8")
        try:
            payload = json.loads(raw)
        except Exception:
            payload = {"error": raw}
        return e.code, payload


def invoice_exists(base_url: str, apikey: str, invoice_number: str) -> bool:
    safe = parse.quote(invoice_number, safe="")
    status, data = supabase_request(
        "GET",
        f"{base_url}/rest/v1/invoices?select=id&invoice_number=eq.{safe}&limit=1",
        apikey,
    )
    return status < 400 and isinstance(data, list) and len(data) > 0


def make_unique_invoice_number(base_url: str, apikey: str, original: str) -> str:
    # Keep original when available; otherwise append -A, -B, ... until unique.
    if not invoice_exists(base_url, apikey, original):
        return original
    suffix_index = 0
    while suffix_index < 200:
        suffix = chr(ord("A") + (suffix_index % 26))
        rounds = suffix_index // 26
        candidate = f"{original}-{suffix}" if rounds == 0 else f"{original}-{suffix}{rounds}"
        if not invoice_exists(base_url, apikey, candidate):
            return candidate
        suffix_index += 1
    # Final safe fallback with timestamp-like entropy
    return f"{original}-{datetime.now(UTC).strftime('%H%M%S')}"


def parse_invoice(xlsx_path: Path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=12, values_only=True)]

    client_name = ""
    invoice_number = ""
    issue_date = ""
    grand_total = 0.0
    payment_deadline = ""
    payment_schedule = []

    sec_titles = [
        "ELECTRICAL AND ELECTRONICS COMPONENTS PRICES",
        "PROTOTYPE ITEMS",
        "SERVICE COST",
    ]
    sections = [
        {"title": sec_titles[0], "columns": [{"key": "sn", "label": "S/N"}, {"key": "item", "label": "Item"}, {"key": "qty", "label": "Qnty"}, {"key": "unitPrice", "label": "Unit price"}, {"key": "totalPrice", "label": "Total price"}], "rows": []},
        {"title": sec_titles[1], "columns": [{"key": "sn", "label": "S/N"}, {"key": "item", "label": "Item"}, {"key": "qty", "label": "Qnty"}, {"key": "unitPrice", "label": "Unit price"}, {"key": "totalPrice", "label": "Total price"}], "rows": []},
        {"title": sec_titles[2], "columns": [{"key": "sn", "label": "S/N"}, {"key": "item", "label": "Item"}, {"key": "qty", "label": "Qnty"}, {"key": "amount", "label": "Amount"}], "rows": []},
    ]

    section_idx = -1
    in_schedule = False
    for r in rows:
        c0, c1, c3, c5, c7, c8 = cell_text(r[0]), cell_text(r[1]), cell_text(r[3]), cell_text(r[5]), cell_text(r[7]), cell_text(r[8])
        up0 = c0.upper()

        if c0 == "Bill to:":
            client_name = c1
        if c7 == "Invoice no:":
            invoice_number = c8
        if c7 == "Invoice Date:":
            issue_date = cell_text(r[8])

        if up0 in sec_titles:
            section_idx = sec_titles.index(up0)
            in_schedule = False
            continue

        if up0 == "PAYMENT SCHEDULE":
            in_schedule = True
            section_idx = -1
            continue

        if in_schedule:
            if isinstance(r[0], (int, float)):
                payment_schedule.append(
                    {
                        "phase": str(int(cell_num(r[0]))),
                        "description": cell_text(r[1]),
                        "amountRatio": cell_num(r[3]),
                        "amountToPay": cell_num(r[5]),
                        "deadline": cell_text(r[7]),
                    }
                )
            elif cell_text(r[1]).upper() == "FINAL":
                payment_deadline = cell_text(r[7])
            continue

        if c3.upper() == "GRAND TOTAL":
            grand_total = cell_num(r[7])
            continue

        if section_idx >= 0 and isinstance(r[0], (int, float)):
            sn = str(int(cell_num(r[0])))
            item = cell_text(r[1])
            qty = cell_num(r[3])
            if section_idx < 2:
                unit_price = cell_num(r[5])
                total_price = cell_num(r[7])
                sections[section_idx]["rows"].append(
                    {"sn": sn, "item": item, "qty": f"{qty:g}", "unitPrice": f"{unit_price:g}", "totalPrice": f"{total_price:g}"}
                )
            else:
                amount = cell_num(r[7])
                sections[section_idx]["rows"].append({"sn": sn, "item": item, "qty": f"{qty:g}", "amount": f"{amount:g}"})

    if not invoice_number:
        invoice_number = f"HC-PI-IMPORT-{datetime.now(UTC).strftime('%Y%m%d%H%M%S')}"
    if not issue_date:
        issue_date = datetime.now(UTC).strftime("%Y-%m-%d")
    due_date = payment_schedule[-1]["deadline"] if payment_schedule else (payment_deadline if payment_deadline[:4].isdigit() else "")

    payload = {
        "dashboardScope": "project",
        "invoiceNumber": invoice_number,
        "clientName": client_name,
        "issueDate": issue_date,
        "dueDate": due_date,
        "currency": "TZS",
        "totals": {"subtotal": grand_total, "taxAmount": 0, "grandTotal": grand_total},
        "projectTables": {
            "sections": sections,
            "paymentSchedule": payment_schedule,
            "paymentGrandTotal": f"{grand_total:g}",
            "paymentDeadline": payment_deadline,
            "note": f"Imported from Excel: {xlsx_path.name}",
        },
    }

    now = datetime.now(UTC).isoformat()
    return {
        "invoice_number": invoice_number,
        "client_name": client_name or "Unknown Client",
        "issue_date": issue_date,
        "due_date": due_date or None,
        "currency": "TZS",
        "subtotal": grand_total,
        "tax_amount": 0,
        "grand_total": grand_total,
        "created_by": None,
        "payload": payload,
        "created_at": now,
        "updated_at": now,
    }


def main():
    repo_root = Path(__file__).resolve().parents[1]
    load_env(repo_root / ".env.local")

    supabase_url = os.getenv("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
    apikey = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("NEXT_PUBLIC_SUPABASE_ANON_KEY")
    if not supabase_url or not apikey:
        raise SystemExit("Missing Supabase env vars in .env.local")

    folder = Path(r"J:\My Drive\Financial records 2026\Proforma invoice 2026")
    skip_name = "SOMEBODY PROFOMA INVOICE.xlsx".lower()
    files = sorted(
        [
            p
            for p in folder.iterdir()
            if p.is_file()
            and p.suffix.lower() in (".xlsx", ".xls")
            and not p.name.startswith("~$")
        ],
        key=lambda p: p.name.lower(),
    )

    inserted = []
    skipped = []
    failed = []

    for f in files:
        if f.name.lower() == skip_name:
            skipped.append({"file": f.name, "reason": "excluded_by_request"})
            continue
        try:
            record = parse_invoice(f)
            inv_no = record["invoice_number"]
            unique_inv_no = make_unique_invoice_number(supabase_url, apikey, inv_no)
            if unique_inv_no != inv_no:
                record["invoice_number"] = unique_inv_no
                record["payload"]["invoiceNumber"] = unique_inv_no
            status, data = 0, None
            last_err = None
            for _ in range(3):
                try:
                    status, data = supabase_request("POST", f"{supabase_url}/rest/v1/invoices", apikey, record)
                    if status < 400:
                        last_err = None
                        break
                    last_err = data
                except Exception as e:
                    last_err = str(e)
            if status >= 400 or last_err:
                failed.append({"file": f.name, "invoice_number": inv_no, "error": last_err or data})
                continue
            row = data[0] if isinstance(data, list) and data else {}
            inserted.append(
                {
                    "file": f.name,
                    "id": row.get("id"),
                    "invoice_number": row.get("invoice_number"),
                    "client_name": row.get("client_name"),
                    "renumbered_from": inv_no if row.get("invoice_number") != inv_no else None,
                }
            )
        except Exception as e:
            failed.append({"file": f.name, "error": str(e)})

    print(json.dumps({"inserted_count": len(inserted), "skipped_count": len(skipped), "failed_count": len(failed), "inserted": inserted, "skipped": skipped, "failed": failed}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

