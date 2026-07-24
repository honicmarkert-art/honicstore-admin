3w`"""Import Electronics components.xlsx as a components-only quotation into Supabase invoices table."""
import json
import os
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib import error, parse, request

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("Install openpyxl: pip install openpyxl") from exc

XLSX_PATH = Path(__file__).resolve().parents[1] / "Electronics components.xlsx"
INVOICE_NUMBER = "QTE-2026-45-COMP-ELECTRONICS"
STAMP_SOURCE_INVOICE_NUMBER = "INV-2026-HCIR-0046"
STAMP_PUBLIC_URL = (
    "https://qobobocldfjhdkpjyuuq.supabase.co/storage/v1/object/public/"
    "invoice-assets/invoices/admin/stamp/company-stamp.jpg"
)
CLIENT_NAME = "Electronics Components (Teaching)"
QUOTATION_SCOPE = (
    "Components only — Orodha ya vifaa vinavyohitajika kufundishia kwa vitendo katika somo la Electronics."
)


def load_env(env_path: Path) -> None:
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def supabase_request(method: str, url: str, apikey: str, body=None):
    headers = {"apikey": apikey, "Authorization": f"Bearer {apikey}"}
    data = None
    if body is not None:
        headers["Content-Type"] = "application/json"
        headers["Prefer"] = "return=representation"
        data = json.dumps(body).encode("utf-8")
    req = request.Request(url, method=method, headers=headers, data=data)
    try:
        with request.urlopen(req) as resp:
            raw = resp.read().decode("utf-8")
            return resp.status, json.loads(raw) if raw else None
    except error.HTTPError as e:
        raw = e.read().decode("utf-8")
        try:
            payload = json.loads(raw)
        except Exception:
            payload = {"error": raw}
        return e.code, payload


def as_number(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def format_unit_suffix(unit: str) -> str:
    unit_text = clean_text(unit).upper()
    if not unit_text:
        return ""
    if unit_text == "PC":
        unit_text = "PCS"
    return f"({unit_text})"


def build_line_description(item: str, detail: str, unit: str) -> str:
    item_text = clean_text(item)
    detail_text = clean_text(detail)
    suffix = format_unit_suffix(unit)
    if not detail_text or detail_text.lower() in item_text.lower():
        base = item_text
    else:
        base = f"{item_text} {detail_text}"
    return f"{base} {suffix}".strip() if suffix else base


def parse_table_title(ws) -> str:
    parts: list[str] = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value:
            text = clean_text(value)
            if text:
                parts.append(text)
    return " ".join(parts)


def parse_excel(path: Path) -> tuple[list[dict], float, str]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    table_title = parse_table_title(ws)
    items: list[dict] = []
    for row_idx in range(3, ws.max_row + 1):
        sn = ws.cell(row_idx, 1).value
        item = ws.cell(row_idx, 2).value
        detail = ws.cell(row_idx, 3).value
        unit = ws.cell(row_idx, 4).value
        qty = ws.cell(row_idx, 5).value
        unit_price = ws.cell(row_idx, 6).value
        total_price = ws.cell(row_idx, 7).value

        if sn is None and item is None:
            continue
        label = clean_text(item)
        if not label:
            continue
        if label.upper() in {"TOTAL", "GRAND TOTAL", "SUBTOTAL"}:
            continue

        quantity = as_number(qty)
        price = as_number(unit_price)
        line_total = as_number(total_price)
        if line_total <= 0 and quantity > 0 and price > 0:
            line_total = quantity * price
        if quantity <= 0 and line_total <= 0:
            continue

        items.append(
            {
                "id": f"comp-{int(sn) if sn is not None and str(sn).isdigit() else row_idx}",
                "description": build_line_description(label, clean_text(detail), clean_text(unit)),
                "quantity": int(quantity) if quantity == int(quantity) else quantity,
                "unitPrice": price,
                "lineTotal": line_total,
            }
        )

    grand_total = sum(row["lineTotal"] for row in items)
    return items, grand_total, table_title


def ensure_unique_invoice_number(base_url: str, apikey: str, invoice_number: str) -> str:
    safe = parse.quote(invoice_number, safe="")
    status, data = supabase_request(
        "GET", f"{base_url}/rest/v1/invoices?select=id&invoice_number=eq.{safe}&limit=1", apikey
    )
    if status >= 400:
        return invoice_number
    if isinstance(data, list) and data:
        suffix = datetime.now().strftime("%H%M%S")
        return f"{invoice_number}-{suffix}"
    return invoice_number


def build_record(items: list[dict], grand_total: float, today: str, due_date: str, invoice_number: str, table_title: str) -> dict:
    payload = {
        "invoiceNumber": invoice_number,
        "documentKind": "quotation",
        "quotationCategory": "components",
        "dashboardScope": "main",
        "issueDate": today,
        "dueDate": due_date,
        "currency": "TZS",
        "taxRate": 0,
        "discount": 0,
        "clientName": CLIENT_NAME,
        "clientEmail": "",
        "clientPhone": "",
        "clientAddress": "",
        "fromName": "Honic Company Limited",
        "fromEmail": "sales@honiccompanystore.com",
        "fromPhone": "+255 763 818138 / +255 786 957 939",
        "companyWebsite": "www.honiccompanystore.com",
        "companyTagline": "ONLINE RETAIL",
        "signerName": "Authorized Signatory",
        "signerTitle": "Administrator",
        "footerPhone": "+255 786 957 939",
        "footerEmail": "support@honiccompany.com",
        "footerAddress": "Dar es Salaam, Tanzania",
        "thankYouLine": "Thank you for considering our proposal.",
        "termsText": (
            "This document is a quotation for electronic components only — not a tax invoice.\n\n"
            "This quotation is valid for 30 days from the issue date. Prices are estimates and subject to availability. "
            "Written acceptance is required before supply begins."
        ),
        "quotationScope": QUOTATION_SCOPE,
        "itemsTableTitle": table_title,
        "referenceNumber": "",
        "items": [
            {
                "id": it["id"],
                "description": it["description"],
                "quantity": it["quantity"],
                "unitPrice": it["unitPrice"],
            }
            for it in items
        ],
        "paymentMethods": [],
        "totals": {"subtotal": grand_total, "taxAmount": 0, "grandTotal": grand_total},
        "importSource": str(XLSX_PATH.name),
    }
    return {
        "invoice_number": invoice_number,
        "client_name": CLIENT_NAME,
        "issue_date": today,
        "due_date": due_date,
        "currency": "TZS",
        "subtotal": grand_total,
        "tax_amount": 0,
        "grand_total": grand_total,
        "created_by": None,
        "payload": payload,
    }


def find_existing_id(base_url: str, apikey: str, invoice_number: str) -> str | None:
    row = find_existing_by_number(base_url, apikey, invoice_number)
    return str(row.get("id") or "") or None if row else None


def find_existing_by_number(base_url: str, apikey: str, invoice_number: str) -> dict | None:
    safe = parse.quote(invoice_number, safe="")
    status, data = supabase_request(
        "GET", f"{base_url}/rest/v1/invoices?select=id,invoice_number,payload&invoice_number=eq.{safe}&limit=1", apikey
    )
    if status >= 400 or not isinstance(data, list) or not data:
        return None
    return data[0]


def main():
    if not XLSX_PATH.exists():
        raise SystemExit(f"Excel file not found: {XLSX_PATH}")

    repo_root = Path(__file__).resolve().parents[1]
    load_env(repo_root / ".env.local")

    supabase_url = os.getenv("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
    apikey = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("NEXT_PUBLIC_SUPABASE_ANON_KEY")
    if not supabase_url or not apikey:
        raise SystemExit("Missing Supabase env vars in honicstore-admin/.env.local")

    items, grand_total, table_title = parse_excel(XLSX_PATH)
    if not items:
        raise SystemExit("No line items parsed from Excel")

    today = datetime.now().strftime("%Y-%m-%d")
    due_date = (datetime.now().date() + timedelta(days=30)).strftime("%Y-%m-%d")
    existing_id = find_existing_id(supabase_url, apikey, INVOICE_NUMBER)
    invoice_number = INVOICE_NUMBER if existing_id else ensure_unique_invoice_number(supabase_url, apikey, INVOICE_NUMBER)
    record = build_record(items, grand_total, today, due_date, invoice_number, table_title)

    record["payload"]["stampImage"] = STAMP_PUBLIC_URL
    now = datetime.now(timezone.utc).isoformat()
    record["updated_at"] = now

    if existing_id:
        patch = {k: v for k, v in record.items() if k != "created_at"}
        status, data = supabase_request(
            "PATCH", f"{supabase_url}/rest/v1/invoices?id=eq.{existing_id}", apikey, patch
        )
        action = "updated"
    else:
        record["created_at"] = now
        status, data = supabase_request("POST", f"{supabase_url}/rest/v1/invoices", apikey, record)
        action = "inserted"

    if status >= 400:
        raise SystemExit(f"Save failed ({status}): {json.dumps(data, ensure_ascii=False)}")

    first = data[0] if isinstance(data, list) and data else {}
    print(
        json.dumps(
            {
                action: True,
                "id": first.get("id", existing_id),
                "invoice_number": first.get("invoice_number", invoice_number),
                "client_name": first.get("client_name"),
                "item_count": len(items),
                "grand_total": first.get("grand_total"),
                "documentKind": "quotation",
                "quotationCategory": "components",
                "sample_descriptions": [it["description"] for it in items[57:60]],
                "items_table_title": table_title,
                "issue_date": today,
                "due_date": due_date,
                "edit_url": f"/dashboard/invoices?invoiceId={first.get('id', existing_id)}&mode=edit",
                "list_url": "/dashboard/invoices/list?documentKind=quotation",
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
